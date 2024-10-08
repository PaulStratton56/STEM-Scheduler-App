from openpyxl import load_workbook
from datetime import datetime, timedelta, date
import json

SCHEDULE_CELLS = 21

EMPTY_SCHEDULE = {
    'Monday' : [], 
    'Tuesday' : [], 
    'Wednesday' : [], 
    'Thursday' : [], 
    'Friday' : []
                }

class Schedule():

    def __init__(self, workbookName: str):
        self.workbook = load_workbook(workbookName)
        self.data = {}

        for sheet in self.workbook.worksheets:
            
            tutors = self.__getTutors(sheet)

            self.__fillTimes(sheet, tutors)

            tutors = self.__getCleanTutors(tutors)

            self.__createShifts(tutors)

            self.data[sheet.title] = tutors

    def __getTutors(self, sheet) -> dict:
        tutors = {}

        for row in sheet.iter_rows(min_row=1, values_only=True):
                tutorName = row[0]
                if (tutorName # exists
                    and tutorName != 'ILA'
                    and tutorName not in tutors.keys()):
                    tutors[tutorName] = {
                        'Monday' : [], 
                        'Tuesday' : [], 
                        'Wednesday' : [], 
                        'Thursday' : [], 
                        'Friday' : []
                        }
        return tutors

    def __fillTimes(self, sheet, tutors) -> dict:
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[1] in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
                day = row[1]
            tutorName = row[0]
            if (tutorName
                and tutorName in tutors.keys()):
                for cell in range(1, SCHEDULE_CELLS):
                    if len(str(row[cell])) == 2:
                        tutors[tutorName][day].append(sheet[chr(65+cell) + str(2)].value)
        
        return tutors

    def __getCleanTutors(self, tutors) -> dict:
        # Remove empty tutors,
        # Or days with empty times.
        cleanTutors = {}
        for tutor, days in tutors.items():
            if (days != EMPTY_SCHEDULE):
                relevantDays = {}
                for day, times in days.items():
                    if times != []:
                        relevantDays[day] = days[day]
                cleanTutors[tutor] = relevantDays

        return cleanTutors
            
    def __createShifts(self, tutors) -> dict:
        for tutor, days in tutors.items():
            for day, times in days.items():  
                intervals = []
                startInterval = 0
                endInterval = startInterval + 1
                while endInterval < len(times):
                    hourDifference = times[endInterval].hour - times[startInterval].hour
                    minuteDifference = times[endInterval].minute - times[startInterval].minute
                    if (hourDifference * 60 + minuteDifference !=
                        (endInterval - startInterval) * 30):
                        intervals.append([times[startInterval],times[endInterval-1]])
                        startInterval = endInterval
                    endInterval += 1
                endTime = datetime.combine(date=date.fromisoformat('2003-10-11'), time=times[endInterval-1])
                endTime += timedelta(minutes=30)
                intervals.append([times[startInterval],endTime])

                tutors[tutor][day] = intervals
        
        return tutors
    
    def printSchedule(self) -> None:
        for subject, tutors in self.data.items():
            print(f"=============== {subject} ===============\n")
            for tutorName, days in tutors.items():
                print(f"{tutorName}: ")
                for day, hours in days.items():
                    print(f"> {day[:3]}: | ", end="")

                    for interval in hours:
                        if (len(interval) == 2):
                            print(f"{interval[0].strftime('%I:%M %p')} - {interval[1].strftime('%I:%M %p')} | ", end="")
                    print()
                print()

    def getJSON(self) -> str:
        dataDict = {}
        for subject, tutors in self.data.items():
            jsonSubject = subject.split('(')[1][:-1]
            tutorDict = {}
            for tutor, days in tutors.items():
                dayDict = {}
                for day, times in days.items():
                    shiftDict = {}
                    for shiftIndex, shift in enumerate(times):
                        shiftStr = shift[0].strftime('%I:%M %p') + '-' + shift[1].strftime('%I:%M %p')
                        shiftDict[shiftIndex] = shiftStr
                    dayDict[day] = shiftDict
                tutorDict[tutor] = dayDict
            dataDict[jsonSubject] = tutorDict
        return json.dumps(dataDict)

if __name__ == '__main__':
    schedule = Schedule("schedules/STEMTutoringFA24Schedule.xlsx")

    # print(schedule.getJSON())
    schedule.printSchedule()
    # print(schedule.data)
    # print(json.dumps(schedule.data))
