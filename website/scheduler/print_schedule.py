import datetime
import openpyxl

# Load an Excel workbook
wb = openpyxl.load_workbook("STEMTutoringFA24Schedule.xlsx")

# Access a specific sheet
sheet = wb["FA 2024 (Math)"]
# sheet = wb["FA 2024 (Chem)"]
# sheet = wb["FA 2024 (Phys)"]
# sheet = wb["FA 2024 (ENGR)"]
# sheet = wb["FA 2024 (CSE)"]

# Access a cell value
cell_value = sheet["A1"].value

# Tutor object
class Tutor:
    def __init__(self, name, schedule):
        self.name = name
        self.schedule = schedule
        
tutor_list = []
for row in sheet.iter_rows(min_row=1, values_only=True):
    if row[0] and row[0] != 'ILA' and row[0] not in (tutor_list[i].name for i in range(len(tutor_list))):
        tutor_list.append(Tutor(row[0], {
        'Monday' : [], 
        'Tuesday' : [], 
        'Wednesday' : [], 
        'Thursday' : [], 
        'Friday' : []
    }))
        
tutor_names = []
for tutor in tutor_list:
    tutor_names.append(tutor.name)
    
tutor_dict = dict(zip(tutor_names, tutor_list))

day_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# print(sheet.dimensions)
# print(sheet['A' + str(1)].value)

def loop_through_rows(sheet):
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[1] in day_list:
            day = row[1]
        if row[0] in (tutor_list[i].name for i in range(len(tutor_list))):
            tutor_name = row[0]
            for i in range(21):
                if len(str(row[i])) == 2:
                    # print(str(sheet[chr(65+i) + str(2)].value) + row[i])
                    tutor_dict[tutor_name].schedule[day].append(sheet[chr(65+i) + str(2)].value)
                    
                
loop_through_rows(sheet)

def print_day(time_list):
    
    first_times = []
    last_times = []
    
    for time in time_list:
        if time == None:
            time_list.remove(time)
    
    for i in range(len(time_list)):
        # if time_list[i] == None:
        #     time_list.remove(None)
        #     if i + 1 == len(time_list):
        #         break
        #     else:
        #         continue
        
        if (time_list[i-1].hour == time_list[i].hour and time_list[i-1].minute + 30 == time_list[i].minute) or (time_list[i-1].hour + 1 == time_list[i].hour and time_list[i-1].minute != time_list[i].minute):
            # print(str(time_list[i]) + " next to previous")
            last_times.append(time_list[i])
            continue
        else:
            # print(str(time_list[i]))
            first_times.append(time_list[i])
            # last_times.append(time_list[i-1])
            
    not_last_times = []
    if last_times:
        old_time = last_times[-1]
    for time in last_times:
        
        if (time.hour == old_time.hour and old_time.minute + 30 == time.minute) or (old_time.hour + 1 == time.hour and old_time.minute != time.minute):
            not_last_times.append(old_time)
        old_time = time
    
    for time in not_last_times:
        last_times.remove(time)
        
    final_last_times = []
    for time in last_times:
        if time.minute == 30:
            new_time = datetime.time(hour=(time.hour+1)%12, minute=0) if time.hour+1 != 12 else datetime.time(hour=(time.hour+1), minute=0)
        else: 
            new_time = datetime.time(hour=(time.hour)%12, minute=30) if time.hour != 12 else datetime.time(hour=time.hour, minute=30)
        final_last_times.append(new_time)
    
    final_first_times = []
    for time in first_times:
        new_time = datetime.time(hour=(time.hour)%12, minute=time.minute) if time.hour != 12 else datetime.time(hour=time.hour, minute=time.minute)
        final_first_times.append(new_time)
    
    for i in range(len(first_times)):
        # print(first_times[i].hour / 12)
        # print(last_times[i].hour / 12)
        if first_times[i].hour / 12 >= 1:
            f_am_pm = 'pm'
        else:
            f_am_pm = 'am'
        if last_times[i].hour / 12 >= 1:
            l_am_pm = 'pm'
        elif last_times[i].hour == 11 and last_times[i].minute == 30:
            l_am_pm = 'pm'
        else:
            l_am_pm = 'am'
        print(str(final_first_times[i].strftime("%H:%M")) + " " + f_am_pm + " - " + str(final_last_times[i].strftime("%H:%M")) + " " + l_am_pm)
        
    # print(first_times)
    # print(last_times)

def print_schedule(tutor):
    
    print("\n" + tutor.name + "\n")
    
    for day in tutor.schedule:
        print(day)
        print_day(tutor.schedule[day])
        print("")


for key in tutor_dict:
    print("________________________")
    print_schedule(tutor_dict[key])
    
    
